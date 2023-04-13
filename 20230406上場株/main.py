import tkinter as tk
import pyautogui
import pandas as pd
import pandas_datareader as pdr
import datetime
from dateutil.relativedelta import relativedelta
import openpyxl

import glob
import PyPDF2

from pdfrw import PdfReader
from pdfrw.buildxobj import pagexobj
from pdfrw.toreportlab import makerl
from reportlab.pdfgen import canvas
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.pagesizes import A4, landscape, portrait
from reportlab.lib.units import mm


root = tk.Tk()
root.geometry("930x930+10+10")
root.title("あげあげくん")
root.resizable(width = False, height= False)
root.iconbitmap(default="ageage.ico")

x = 1
y = 20
NumEnt = [[0 for k in range(y)] for l in range(x)]
print(NumEnt)

for i in range(x):
    for j in range(y):
        sv = tk.StringVar()
        NumEnt[i][j] = tk.Entry(root, width = 4,font = ("",30),textvariable = sv)
        def validation(before_word, after_word):
            return ((after_word.isdecimal()) and (len(after_word)<=4)) or (len(after_word) == 0)
        vcmd = (NumEnt[i][j].register(validation), '%s', '%P')
        NumEnt[i][j].configure(validate='key', vcmd=vcmd)
        NumEnt[i][j].place(x = i*100+30, y = j*45+30)

        def tabkey(event):
            pyautogui.press("tab")
        NumEnt[i][j].bind('<Return>', tabkey)

tk.Label(root, text = "あげあげくん", font = ("",50)).place(x = 400, y = 30)

tk.Label(root, text = "相続発生日", font = ("",20)).place(x = 400, y = 130)

sv2 = tk.StringVar()
Dday = tk.Entry(root, width = 8, font = ("",20),textvariable = sv2)
def validation(before_word, after_word):
    return ((after_word.isdecimal()) and (len(after_word)<=8)) or (len(after_word) == 0)
vcmd2 = (Dday.register(validation), '%s', '%P')
Dday.configure(validate='key', vcmd=vcmd2)
Dday.place(x = 600, y = 130)

def Main():
    global x, y ,NumEnt
    AllCode = []
    for i in range(x):
        for j in range(y):
            p = NumEnt[i][j]
            try:
                q = int(p.get())
                if q < 10000 and q > 1000:
                    AllCode.append(q)
                    pass
            except:
                pass
            
    print(AllCode)
    DD = Dday.get()
    print(DD)
    dte = datetime.datetime.strptime(DD, '%Y%m%d')
    print(dte)
    # today = dte.date()
    # print(today)
    Start = dte-datetime.timedelta(days = 10)
    print(Start)
    End = dte+datetime.timedelta(days = 10)
    
    ToMonth = dte.date()
    LMonth = ToMonth-relativedelta(months=1)
    LLMonth = LMonth-relativedelta(months=1)
    ToMonth = ToMonth.strftime('%Y%m')
    LMonth = LMonth.strftime('%Y%m')
    LLMonth = LLMonth.strftime('%Y%m')
    print(AllCode, DD, Start, End, ToMonth, LMonth, LLMonth)
    wb = openpyxl.load_workbook('Base.xlsx')
    ws = wb.worksheets[0]
    for n, number in enumerate(AllCode):
        path = "monthexc/" + ToMonth + ".xlsx"
        dff = pd.read_excel(path)
        x = dff.iloc[:,1]
        y = dff.iloc[:,2]
        z = dff.iloc[:,11]
        # 必要な部分だけ抽出結合
        dff1 = pd.concat([x,y,z], axis=1)
        # https://note.nkmk.me/python-pandas-concat/
        dff1.columns=["コード", "名前", "終値平均"]
        df3 =dff1[dff1['コード'] == int(number)]
        print(df3)
        print("aruyo-")


        Code = str(number) + ".jp"
        df = pdr.DataReader(Code, data_source = "stooq", start = Start, end = End)
        if dte in df.index:
            print('あるよ')
            TODAY_V = df.at[dte,"Close"]
            print(TODAY_V)
        else:
            day1 = dte-datetime.timedelta(days = 1)
            day2 = dte+datetime.timedelta(days = 1)
            if day1 in df.index and day2 in df.index:
                TODAY_V = (df.at[day1, "Close"] + df.at[day2,"Close"])/2
            elif day1 in df.index:
                TODAY_V = df.at[day1,"Close"]
            elif day2 in df.index:
                TODAY_V = df.at[day2,"Close"]
            else:
                day1 = day1-datetime.timedelta(days = 1)
                day2 = day2+datetime.timedelta(days = 1)
                if day1 in df.index and day2 in df.index:
                    TODAY_V = (df.at[day1, "Close"] + df.at[day2,"Close"])/2
                elif day1 in df.index:
                    TODAY_V = df.at[day1,"Close"]
                elif day2 in df.index:
                    TODAY_V = df.at[day2,"Close"]
                else:
                    day1 = day1-datetime.timedelta(days = 1)
                    day2 = day2+datetime.timedelta(days = 1)
                    if day1 in df.index and day2 in df.index:
                        TODAY_V = (df.at[day1, "Close"] + df.at[day2,"Close"])/2
                    elif day1 in df.index:
                        TODAY_V = df.at[day1,"Close"]
                    elif day2 in df.index:
                        TODAY_V = df.at[day2,"Close"]
                    else:
                        day1 = day1-datetime.timedelta(days = 1)
                        day2 = day2+datetime.timedelta(days = 1)
                        if day1 in df.index and day2 in df.index:
                            TODAY_V = (df.at[day1, "Close"] + df.at[day2,"Close"])/2
                        elif day1 in df.index:
                            TODAY_V = df.at[day1,"Close"]
                        elif day2 in df.index:
                            TODAY_V = df.at[day2,"Close"]
                        else:
                            day1 = day1-datetime.timedelta(days = 1)
                            day2 = day2+datetime.timedelta(days = 1)
                            if day1 in df.index and day2 in df.index:
                                TODAY_V = (df.at[day1, "Close"] + df.at[day2,"Close"])/2
                            elif day1 in df.index:
                                TODAY_V = df.at[day1,"Close"]
                            elif day2 in df.index:
                                TODAY_V = df.at[day2,"Close"]
                            else:
                                print("おかしいね")
                                break
        print(TODAY_V)
        print(number)
        print(n)
        TLLL = [LLMonth, LMonth, ToMonth]
        MValue = []
        for i in TLLL:
            path1 = "monthexc/" + i + ".xlsx"
            dff = pd.read_excel(path1)
            x = dff.iloc[:,1]
            y = dff.iloc[:,2]
            z = dff.iloc[:,11]
            # 必要な部分だけ抽出結合
            dff1 = pd.concat([x,y,z], axis=1)
            # https://note.nkmk.me/python-pandas-concat/
            dff1.columns=["コード", "名前", "終値平均"]
            # df2 = dff1.dropna(how='any')
            # https://note.nkmk.me/python-pandas-nan-dropna/
            # print(df2)
            print(dff1)
            # 検索
            df3 =dff1[dff1['コード'] == int(number)]
            print(df3)
            MValue.append(int(df3.iat[0,2]))
        name1 = df3.iat[0,1]
        name = name1[:-5]
        MValue.append(int(TODAY_V))
        Mmin = min(MValue)
        print("いれました！")
        ws.cell(row = 5+n,column = 2).value = str(number)
        ws.cell(row = 5+n,column = 3).value = str(name)
        ws.cell(row = 5+n,column = 4).value = str(int(TODAY_V))
        ws.cell(row = 5+n,column = 5).value = str(int(MValue[2]))
        ws.cell(row = 5+n,column = 6).value = str(int(MValue[1]))
        ws.cell(row = 5+n,column = 7).value = str(int(MValue[0]))
        ws.cell(row = 5+n,column = 8).value = str(Mmin)
        #別名で保存
    wb.save('完成品/上場株評価.xlsx')
    root.destroy()



MainBTN = tk.Button(root, text = "実行", command = Main, font = ("",15))
MainBTN.place(x = 750, y = 130)


root.mainloop()
